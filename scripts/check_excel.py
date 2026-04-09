"""Check ALL rows in Excel for the 5 problematic products."""
from pathlib import Path
import argparse
import openpyxl, sys, io, unicodedata

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

DEFAULT_EXCEL = Path(__file__).resolve().parents[1] / "Shopify_Publicaciones_Exportación_0404221023.xlsx"


def parse_args():
    parser = argparse.ArgumentParser(description="Check all rows in an Excel workbook for the 5 problematic products.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help=f"Path to the Excel workbook (default: {DEFAULT_EXCEL})")
    return parser.parse_args()


def normalize(text):
    text = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in text if not unicodedata.combining(c)).lower().strip()

def main():
    args = parse_args()
    excel_path = Path(args.excel).expanduser()
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # The 5 problematic product titles (keywords to search)
    problem_keywords = [
        "vaseline glazed",
        "hershey",
        "vitamina c 1000",
        "zma",
    ]

    print("=== ALL ROWS IN EXCEL ===")
    print(f"Workbook: {excel_path}")
    print(f"Total rows: {ws.max_row - 1}")
    print()

    # Show ALL rows with their IDs
    for row in ws.iter_rows(min_row=2, values_only=True):
        pub_id = str(row[1]).strip() if row[1] else ""
        title = str(row[2]).strip() if row[2] else ""

        if not title:
            continue

        # Check if this is one of the problematic products
        is_problem = any(kw in title.lower() for kw in problem_keywords)

        if is_problem:
            marker = " <<< PROBLEM PRODUCT"
        else:
            marker = ""

        print(f"ID: {pub_id} | {title[:80]}{marker}")


if __name__ == "__main__":
    main()
